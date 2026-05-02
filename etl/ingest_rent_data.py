"""One-time ingestion of average rent prices (NIS) from a manually
prepared Excel workbook into the cbs_rent_prices table.

The workbook is itself a manual extraction from CBS quarterly rent
publications (PDFs) — CBS does NOT expose this data via any API
surface (we checked — the time-series API, price-index API, and
boards generator all only have rent indices, not absolute NIS rent
levels). The Excel was hand-built from the source PDFs and spot-
verified before ingestion.

Source: etl/data/rent_data_2017_2025.xlsx, sheet `Tidy_Data` (already
in long format — Geography Type, Geography, Room Group, Year, Period,
Average Rent (NIS)).

We only persist QUARTERLY rows. Annual averages from the Excel are
skipped — they can be recomputed from the quarters when needed, and
keeping them in the DB risks divergence (e.g., the 2021 source PDF
omitted annuals so the Excel has gaps that wouldn't exist for a
computed annual). The schema reflects this — there is no `frequency`
column on cbs_rent_prices.

Period mapping (annuals dropped at parse time, before reaching DB):
    'Annual' → SKIPPED
    'Q1'     → time_period = YYYY-01-01
    'Q2'     → time_period = YYYY-04-01
    'Q3'     → time_period = YYYY-07-01
    'Q4'     → time_period = YYYY-10-01

Geography mapping:
    Type 'National' (geo='סך הכל')      → ('national', 'national')
    Type 'District' (Hebrew + 'מחוז ')   → ('district', english_key)
    Type 'City'                           → ('city', Hebrew name kept)

Room group normalization: the canonical set is {'all', '1-2',
'2.5-3', '3.5-4', '4.5-6'}. We additionally handle two known data-
hygiene issues:

  - Unicode dash variants (en-dash U+2013, em-dash U+2014, figure-
    dash U+2012, minus-sign U+2212, non-breaking hyphen U+2011) are
    folded to ASCII hyphen-minus, so '4.5–6' → '4.5-6'. Generic; not
    triggered by anything we've seen yet, but cheap insurance for
    future PDF updates that might preserve typographic dashes.
  - Specific known PDF-tokenization split-error: the cell '4.5-6'
    appears in the 2017-2019 Be'er Sheva block as '- 6' (the prefix
    '4.5' having leaked into a value cell — see the sanity floor
    below for that side of the issue). We map '- 6' (with surrounding
    whitespace tolerated) → '4.5-6'. Logged when triggered so we can
    audit if/when the Excel is rebuilt.

source_pdf attribution comes from the workbook's README sheet, which
lists which PDF supplied each year/period combination. The script
parses that README at runtime so the mapping stays in sync if the
extraction is rebuilt — no hard-coded year→PDF dictionary in code.

Sanity floor: any rent value < SANITY_FLOOR_NIS is rejected as a
data-quality error and logged. Average rent for a Be'er Sheva
apartment is in the thousands of NIS; the floor catches values like
4.5 that crept in via tokenization splits in the source PDF
extraction. We drop the row rather than reverse-engineering from
the annual — explicit gap is preferable to fabricated data.

Idempotent: ON CONFLICT (geography_type, geography, room_group,
time_period) DO UPDATE on each upsert. Safe to re-run when the
Excel is regenerated.

Rows with NaN rent (notably 2021 annuals — that year's annuals
were missing from the source PDF) are skipped silently. Logged at
end with a count.
"""
from __future__ import annotations

import logging
import os
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from supabase import Client, create_client

EXCEL_PATH = Path(__file__).parent / 'data' / 'rent_data_2017_2025.xlsx'
TABLE = 'cbs_rent_prices'
BATCH_SIZE = 500

# Below this threshold (in NIS) a value is treated as a data-quality
# error rather than a real rent reading. Real average rents for any
# city/room-group combination are in the low thousands; values like
# 4.5 are leaked label fragments from PDF tokenization splits.
SANITY_FLOOR_NIS = 100.0

# Unicode dash characters that should be normalized to ASCII hyphen-
# minus before room-group matching. Folded together so a `'4.5–6'`
# cell (with en-dash) becomes `'4.5-6'` and matches the canonical
# value, etc.
DASH_VARIANTS = (
    '‐',  # HYPHEN
    '‑',  # NON-BREAKING HYPHEN
    '‒',  # FIGURE DASH
    '–',  # EN DASH
    '—',  # EM DASH
    '−',  # MINUS SIGN
)

# Hebrew district name → English snake_case key (matching the
# convention used elsewhere in the project; see DISTRICT_DB_KEY in
# src/data/seriesRegistry.ts).
DISTRICT_HE_TO_EN = {
    'מחוז דרום':     'south',
    'מחוז חיפה':     'haifa',
    'מחוז ירושלים':  'jerusalem',
    'מחוז צפון':     'north',
    'מחוז מרכז':     'center',
    'מחוז תל אביב':  'tel_aviv',
}

# Maps Tidy_Data's Period column to month-of-year-start. Annual rows
# are filtered out before this map is consulted (we persist quarterly
# only — see module docstring for rationale).
QUARTER_MONTH = {'Q1': 1, 'Q2': 4, 'Q3': 7, 'Q4': 10}

logging.basicConfig(
    level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s',
)
log = logging.getLogger('ingest_rent')


def parse_source_pdf_map(wb: openpyxl.Workbook) -> dict[tuple[int, str], str]:
    """Build a (year, period) → source_pdf_filename mapping from the
    README sheet's "Source PDF / Years/periods used" rows.

    Each README row reads like:
      ('2019-2017.pdf', '2017-2019 annual and quarterly data')
      ('2022-2021.pdf', '2022 annual and quarterly data; 2021 quarterly data')

    The right-hand description is parsed for year ranges and the
    period qualifier ("annual", "quarterly", or both). When a year
    appears with both qualifiers, all 5 (Annual + Q1-Q4) entries
    point to the same PDF; with only one qualifier, only those
    entries are mapped. PDF rows whose description doesn't match
    the expected pattern are logged and skipped.

    Years/PDFs not present here just get source_pdf=None — the
    column is nullable, so this is fine."""
    if 'README' not in wb.sheetnames:
        log.warning('README sheet missing; source_pdf will be NULL on all rows')
        return {}
    ws = wb['README']
    out: dict[tuple[int, str], str] = {}
    # README header row "Source PDF | Years/periods used" appears
    # somewhere; everything beneath it is one PDF per row. Scan the
    # whole sheet — README is small.
    started = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        a, b = (row[0], row[1]) if len(row) >= 2 else (row[0], None)
        if a is None:
            continue
        if isinstance(a, str) and a.strip() == 'Source PDF':
            started = True
            continue
        if not started:
            continue
        if not isinstance(a, str) or not a.lower().endswith('.pdf'):
            continue
        if not isinstance(b, str):
            continue
        # Extract years (4-digit numbers) and qualifiers from the
        # description. Examples we handle:
        #   "2017-2019 annual and quarterly data"
        #   "2020 annual and quarterly data; duplicate 2019 ..."
        #   "2022 annual and quarterly data; 2021 quarterly data"
        # The semicolon-delimited form is parsed by clause so each
        # year gets its own (qualifiers).
        for clause in re.split(r';', b):
            yrs = [int(y) for y in re.findall(r'(?<!\d)(\d{4})(?!\d)', clause)]
            if not yrs:
                continue
            # Year range "2017-2019" → expand to [2017, 2018, 2019]
            if len(yrs) == 2 and yrs[1] - yrs[0] in (1, 2, 3):
                yrs = list(range(yrs[0], yrs[1] + 1))
            cl = clause.lower()
            has_annual = 'annual' in cl
            has_quart  = 'quarterly' in cl or 'quarter' in cl
            # If neither qualifier appears, assume both (lenient).
            if not has_annual and not has_quart:
                has_annual = has_quart = True
            for y in yrs:
                if has_annual:
                    out[(y, 'Annual')] = a
                if has_quart:
                    for q in ('Q1', 'Q2', 'Q3', 'Q4'):
                        out[(y, q)] = a
    log.info('parsed README: %d (year, period) → PDF entries', len(out))
    return out


def map_geography(gtype: str, geo: str) -> tuple[str, str] | None:
    """Translate Excel's (Geography Type, Geography) → DB
    (geography_type, geography) pair. Returns None for unknown
    inputs so the caller can log + skip."""
    if gtype == 'National':
        # Excel uses 'סך הכל' for the national row regardless of
        # locale convention; collapse to a single 'national' key.
        return ('national', 'national')
    if gtype == 'District':
        en = DISTRICT_HE_TO_EN.get(geo)
        if en is None:
            return None
        return ('district', en)
    if gtype == 'City':
        # Cities stay in Hebrew — no English mapping table exists
        # for them in the project, and the chart UI is RTL-friendly.
        return ('city', geo)
    return None


def normalize_room_group(rg: str) -> tuple[str | None, str | None]:
    """Return (canonical_room_group, normalization_note).

    Performs in order:
      1. Strip whitespace.
      2. Replace Unicode dash variants (en-dash, em-dash, etc.) with
         ASCII hyphen-minus.
      3. Map 'All' → 'all'.
      4. Special-case the known PDF-extraction split error '- 6'
         (hyphen-space-six) → '4.5-6'. The leading '4.5' having
         leaked into a value cell is handled separately by the
         sanity floor on the rent value.
      5. Validate against the canonical room-group set.

    Returns (None, None) if the input is empty / non-string.
    Returns (canonical, None) if the input was already canonical.
    Returns (canonical, note) if normalization was applied — the
    caller logs `note` so audits can trace which rows got fixed.
    Returns (None, note) if the input doesn't match anything even
    after normalization (caller skips with the note in the warning).
    """
    if not isinstance(rg, str):
        return None, None
    original = rg
    cleaned = rg.strip()
    for d in DASH_VARIANTS:
        cleaned = cleaned.replace(d, '-')
    if cleaned == 'All':
        return 'all', None
    if cleaned in ('1-2', '2.5-3', '3.5-4', '4.5-6'):
        # If we had to do anything (whitespace strip / dash fold) to
        # get here, surface that as a note so it's auditable.
        if cleaned != original:
            return cleaned, f'normalized {original!r} → {cleaned!r}'
        return cleaned, None
    if cleaned == '- 6':
        # Known PDF-tokenization split error in the source data.
        # Documented at module level.
        return '4.5-6', f'recovered {original!r} → "4.5-6" (PDF split-error)'
    return None, f'unrecognized room group {original!r}'


def build_rows(wb: openpyxl.Workbook) -> list[dict]:
    """Iterate Tidy_Data, transform, and return a list of dicts
    ready for `supabase.upsert`. Logs counts of skipped rows by
    reason at the end so the operator can verify nothing important
    was lost."""
    if 'Tidy_Data' not in wb.sheetnames:
        raise RuntimeError(
            'Tidy_Data sheet missing from workbook — '
            'cannot ingest. Re-export rent_data_2017_2025.xlsx.'
        )
    ws = wb['Tidy_Data']
    pdf_map = parse_source_pdf_map(wb)
    rows: list[dict] = []
    skipped_nan = 0
    skipped_annual = 0
    skipped_below_floor: list[tuple] = []
    skipped_unknown_geo: list[tuple[str, str]] = []
    skipped_unknown_rg: list[tuple[str, str]] = []  # (room_value, note)
    skipped_unknown_period: list[str] = []
    normalized_rg: list[str] = []
    header = None
    for r in ws.iter_rows(min_row=1, values_only=True):
        if header is None:
            header = r
            continue
        gtype, geo, room, year, period, rent = r
        if gtype is None:
            continue  # blank row
        # Drop annual rows pre-validation: we persist quarterly only
        # (annuals are recomputed downstream when needed).
        if period == 'Annual':
            skipped_annual += 1
            continue
        if rent is None:
            skipped_nan += 1
            continue
        geo_pair = map_geography(gtype, geo)
        if geo_pair is None:
            skipped_unknown_geo.append((gtype, geo))
            continue
        rg, note = normalize_room_group(room)
        if rg is None:
            skipped_unknown_rg.append((str(room), note or ''))
            continue
        if note:
            normalized_rg.append(note)
        month = QUARTER_MONTH.get(period)
        if month is None:
            skipped_unknown_period.append(period)
            continue
        rent_f = float(rent)
        if rent_f < SANITY_FLOOR_NIS:
            # Real avg rents are in the thousands of NIS. A value
            # this low is a data-quality artifact (e.g., the leaked
            # '4.5' from the '4.5-6' label split). Drop it; better
            # to have an explicit gap than fabricated data.
            skipped_below_floor.append(
                (gtype, geo, room, int(year), period, rent_f),
            )
            continue
        time_period = date(int(year), month, 1)
        rows.append({
            'geography_type': geo_pair[0],
            'geography': geo_pair[1],
            'room_group': rg,
            'time_period': time_period.isoformat(),
            'value': rent_f,
            'source_pdf': pdf_map.get((int(year), period)),
            'is_estimated': False,
        })

    # Deduplicate by the table's unique key. The source Excel
    # contains some legitimate-looking duplicates that only collide
    # after room-group normalization — e.g., 2019 Be'er Sheva has
    # both '4.5-6' rows and the recovered-from-'- 6' rows with
    # identical values. Exact-value duplicates are silently dropped;
    # value-mismatch duplicates are logged loudly so we can audit
    # which side to trust before re-running.
    dedup: dict[tuple, dict] = {}
    exact_dupes = 0
    value_mismatches: list[tuple] = []
    for row in rows:
        key = (
            row['geography_type'], row['geography'],
            row['room_group'], row['time_period'],
        )
        existing = dedup.get(key)
        if existing is None:
            dedup[key] = row
            continue
        if existing['value'] == row['value']:
            exact_dupes += 1
            continue
        # Different values for the same key — keep the first
        # occurrence (chronological by Excel row order) and warn.
        value_mismatches.append((key, existing['value'], row['value']))
    deduped_rows = list(dedup.values())

    log.info(
        'parsed Tidy_Data: %d rows kept after dedup, %d skipped (Annual, '
        'by design), %d skipped (NaN), %d skipped (unknown geo), %d skipped '
        '(unknown room group), %d skipped (unknown period), %d skipped '
        '(below sanity floor of %.0f NIS), %d exact-match duplicates '
        'collapsed',
        len(deduped_rows), skipped_annual, skipped_nan,
        len(skipped_unknown_geo), len(skipped_unknown_rg),
        len(skipped_unknown_period), len(skipped_below_floor),
        SANITY_FLOOR_NIS, exact_dupes,
    )
    if value_mismatches:
        log.warning(
            'value-mismatch duplicates found (%d) — kept FIRST, ignored '
            'the rest. Inspect the source Excel before trusting these:',
            len(value_mismatches),
        )
        for key, kept, ignored in value_mismatches:
            log.warning('  %s : kept=%s, ignored=%s', key, kept, ignored)
    if normalized_rg:
        log.info(
            'room group normalizations applied (%d total):', len(normalized_rg),
        )
        for n in sorted(set(normalized_rg)):
            log.info('  %s', n)
    if skipped_below_floor:
        log.warning('rows dropped by sanity floor (rent < %.0f NIS):',
                    SANITY_FLOOR_NIS)
        for s in skipped_below_floor:
            log.warning('  %s', s)
    if skipped_unknown_geo:
        unique_geo = sorted(set(skipped_unknown_geo))[:10]
        log.warning('unknown geographies (first 10): %s', unique_geo)
    if skipped_unknown_rg:
        log.warning(
            'unknown room groups (with normalization notes): %s',
            sorted(set(skipped_unknown_rg))[:10],
        )
    if skipped_unknown_period:
        log.warning('unknown periods: %s', sorted(set(skipped_unknown_period)))
    return deduped_rows


def upsert_rows(client: Client, rows: list[dict]) -> int:
    if not rows:
        return 0
    sent = 0
    for start in range(0, len(rows), BATCH_SIZE):
        batch = rows[start : start + BATCH_SIZE]
        client.table(TABLE).upsert(
            batch,
            on_conflict='geography_type,geography,room_group,time_period',
        ).execute()
        sent += len(batch)
        log.info('upserted batch %d-%d (%d total)', start, start + len(batch), sent)
    return sent


def main() -> int:
    load_dotenv(Path(__file__).parent / '.env')
    url = os.environ.get('SUPABASE_URL')
    key = os.environ.get('SUPABASE_SERVICE_ROLE_KEY')
    if not url or not key:
        log.error('SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set')
        return 2
    if not EXCEL_PATH.exists():
        log.error('Excel not found at %s', EXCEL_PATH)
        return 2
    log.info('reading %s', EXCEL_PATH)
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    rows = build_rows(wb)
    if not rows:
        log.error('no rows to upsert')
        return 1
    client = create_client(url, key)
    sent = upsert_rows(client, rows)
    log.info('Done. Upserted %d rent rows.', sent)
    return 0


if __name__ == '__main__':
    sys.exit(main())
